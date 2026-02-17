import csv
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, Dict, List


class RelatorioExportService:
    def _as_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value)

    def _as_float(self, value: Any) -> float:
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value))
        except (TypeError, ValueError):
            return 0.0

    def _as_int(self, value: Any) -> int:
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        try:
            return int(float(str(value)))
        except (TypeError, ValueError):
            return 0

    def _as_latin1(self, value: Any) -> str:
        return self._as_text(value).encode("latin-1", errors="replace").decode("latin-1")

    def _build_dimension_rows(self, relatorio: Dict[str, Any]) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        dominios = relatorio.get("dominios") or []
        for dominio in dominios:
            dominio_codigo = self._as_text(dominio.get("codigo"))
            dominio_nome = self._as_text(dominio.get("nome") or "Sem domínio")
            dimensoes = dominio.get("dimensoes") or []
            for dimensao in dimensoes:
                distribuicao = dimensao.get("distribuicao") or {}
                classificacao_raw = dimensao.get("classificacao")
                classificacao = (
                    classificacao_raw.get("value")
                    if isinstance(classificacao_raw, dict) and "value" in classificacao_raw
                    else classificacao_raw
                )

                rows.append(
                    {
                        "dominio_codigo": dominio_codigo,
                        "dominio_nome": dominio_nome,
                        "dimensao": self._as_text(dimensao.get("dimensao") or "Sem dimensão"),
                        "media": round(self._as_float(dimensao.get("media")), 2),
                        "classificacao": self._as_text(classificacao or "intermediario"),
                        "sinal": self._as_text(dimensao.get("sinal") or "risco"),
                        "favoravel": self._as_int(distribuicao.get("favoravel")),
                        "intermediario": self._as_int(distribuicao.get("intermediario")),
                        "risco": self._as_int(distribuicao.get("risco")),
                    }
                )
        return rows

    def _build_filename_base(self, relatorio: Dict[str, Any]) -> str:
        relatorio_id = self._as_text(relatorio.get("id") or "sem_id")
        tipo = self._as_text(relatorio.get("tipoRelatorio") or "relatorio")
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return f"{tipo}_{relatorio_id[:8]}_{timestamp}"

    def build_csv_bytes(self, relatorio: Dict[str, Any]) -> bytes:
        output = StringIO()
        writer = csv.writer(output, delimiter=";")
        metricas = relatorio.get("metricas") or {}

        writer.writerow(["Relatório", self._as_text(relatorio.get("id"))])
        writer.writerow(["Tipo", self._as_text(relatorio.get("tipoRelatorio"))])
        writer.writerow(["Gerado por", self._as_text(relatorio.get("geradoPor"))])
        writer.writerow(["Data de geração", self._as_text(relatorio.get("dataGeracao"))])
        writer.writerow([])
        writer.writerow(["Métrica", "Valor"])
        writer.writerow(["Média de Risco Global", f"{self._as_float(metricas.get('mediaRiscoGlobal')):.2f}"])
        writer.writerow(["Índice de Proteção (%)", f"{self._as_float(metricas.get('indiceProtecao')):.2f}"])
        writer.writerow(["Total de Respondentes", self._as_int(metricas.get("totalRespondentes"))])
        writer.writerow([])
        writer.writerow(
            [
                "Código Domínio",
                "Domínio",
                "Dimensão",
                "Média",
                "Classificação",
                "Sinal",
                "Favorável",
                "Intermediário",
                "Risco",
            ]
        )

        for row in self._build_dimension_rows(relatorio):
            writer.writerow(
                [
                    row["dominio_codigo"],
                    row["dominio_nome"],
                    row["dimensao"],
                    f"{row['media']:.2f}",
                    row["classificacao"],
                    row["sinal"],
                    row["favoravel"],
                    row["intermediario"],
                    row["risco"],
                ]
            )

        writer.writerow([])
        writer.writerow(["Recomendações"])
        recomendacoes = relatorio.get("recomendacoes") or []
        if recomendacoes:
            for index, recomendacao in enumerate(recomendacoes, start=1):
                writer.writerow([f"{index}. {self._as_text(recomendacao)}"])
        else:
            writer.writerow(["Sem recomendações."])

        return output.getvalue().encode("utf-8-sig")

    def build_excel_bytes(self, relatorio: Dict[str, Any]) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise RuntimeError("Dependência 'openpyxl' não instalada para exportação Excel.") from exc

        wb = Workbook()
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        metricas = relatorio.get("metricas") or {}

        ws_resumo.append(["Relatório", self._as_text(relatorio.get("id"))])
        ws_resumo.append(["Tipo", self._as_text(relatorio.get("tipoRelatorio"))])
        ws_resumo.append(["Gerado por", self._as_text(relatorio.get("geradoPor"))])
        ws_resumo.append(["Data de geração", self._as_text(relatorio.get("dataGeracao"))])
        ws_resumo.append([])
        ws_resumo.append(["Métrica", "Valor"])
        ws_resumo.append(["Média de Risco Global", round(self._as_float(metricas.get("mediaRiscoGlobal")), 2)])
        ws_resumo.append(["Índice de Proteção (%)", round(self._as_float(metricas.get("indiceProtecao")), 2)])
        ws_resumo.append(["Total de Respondentes", self._as_int(metricas.get("totalRespondentes"))])

        ws_resumo["A1"].font = Font(bold=True)
        ws_resumo["A6"].font = Font(bold=True)
        ws_resumo.column_dimensions["A"].width = 30
        ws_resumo.column_dimensions["B"].width = 36

        ws_dimensoes = wb.create_sheet("Dominios e Dimensoes")
        ws_dimensoes.append(
            [
                "Código Domínio",
                "Domínio",
                "Dimensão",
                "Média",
                "Classificação",
                "Sinal",
                "Favorável",
                "Intermediário",
                "Risco",
            ]
        )
        for cell in ws_dimensoes[1]:
            cell.font = Font(bold=True)

        for row in self._build_dimension_rows(relatorio):
            ws_dimensoes.append(
                [
                    row["dominio_codigo"],
                    row["dominio_nome"],
                    row["dimensao"],
                    row["media"],
                    row["classificacao"],
                    row["sinal"],
                    row["favoravel"],
                    row["intermediario"],
                    row["risco"],
                ]
            )

        ws_dimensoes.column_dimensions["A"].width = 16
        ws_dimensoes.column_dimensions["B"].width = 28
        ws_dimensoes.column_dimensions["C"].width = 34
        ws_dimensoes.column_dimensions["D"].width = 10
        ws_dimensoes.column_dimensions["E"].width = 16
        ws_dimensoes.column_dimensions["F"].width = 10
        ws_dimensoes.column_dimensions["G"].width = 12
        ws_dimensoes.column_dimensions["H"].width = 14
        ws_dimensoes.column_dimensions["I"].width = 10

        ws_recomendacoes = wb.create_sheet("Recomendacoes")
        ws_recomendacoes.append(["Prioridade", "Recomendação"])
        ws_recomendacoes["A1"].font = Font(bold=True)
        ws_recomendacoes["B1"].font = Font(bold=True)
        recomendacoes = relatorio.get("recomendacoes") or []
        if recomendacoes:
            for index, recomendacao in enumerate(recomendacoes, start=1):
                ws_recomendacoes.append([index, self._as_text(recomendacao)])
        else:
            ws_recomendacoes.append([1, "Sem recomendações."])
        ws_recomendacoes.column_dimensions["A"].width = 12
        ws_recomendacoes.column_dimensions["B"].width = 90

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def build_pdf_bytes(self, relatorio: Dict[str, Any]) -> bytes:
        try:
            from fpdf import FPDF
        except ImportError as exc:
            raise RuntimeError("Dependência 'fpdf2' não instalada para exportação PDF.") from exc

        metricas = relatorio.get("metricas") or {}
        rows = self._build_dimension_rows(relatorio)
        recomendacoes = relatorio.get("recomendacoes") or []

        pdf = FPDF(format="A4")
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.add_page()

        pdf.set_font("Helvetica", "B", 15)
        pdf.cell(0, 9, self._as_latin1("Relatório Consolidado COPSOQ"), new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 6, self._as_latin1(f"ID: {self._as_text(relatorio.get('id'))}"), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, self._as_latin1(f"Tipo: {self._as_text(relatorio.get('tipoRelatorio'))}"), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, self._as_latin1(f"Gerado por: {self._as_text(relatorio.get('geradoPor'))}"), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, self._as_latin1(f"Data: {self._as_text(relatorio.get('dataGeracao'))}"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, self._as_latin1("Métricas principais"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 6, self._as_latin1(f"Média de risco global: {self._as_float(metricas.get('mediaRiscoGlobal')):.2f}"), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, self._as_latin1(f"Índice de proteção: {self._as_float(metricas.get('indiceProtecao')):.2f}%"), new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, self._as_latin1(f"Total de respondentes: {self._as_int(metricas.get('totalRespondentes'))}"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, self._as_latin1("Domínios e dimensões"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=9)
        if rows:
            for row in rows:
                linha = (
                    f"{row['dominio_nome']} | {row['dimensao']} | "
                    f"media={row['media']:.2f} | class={row['classificacao']} | "
                    f"fav={row['favoravel']} int={row['intermediario']} risco={row['risco']}"
                )
                pdf.multi_cell(0, 5, self._as_latin1(linha))
        else:
            pdf.multi_cell(0, 5, self._as_latin1("Sem dimensões no relatório."))
        pdf.ln(1)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, self._as_latin1("Recomendações"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=10)
        if recomendacoes:
            for index, recomendacao in enumerate(recomendacoes, start=1):
                pdf.multi_cell(0, 6, self._as_latin1(f"{index}. {self._as_text(recomendacao)}"))
        else:
            pdf.multi_cell(0, 6, self._as_latin1("Sem recomendações."))

        output = pdf.output(dest="S")
        if isinstance(output, (bytes, bytearray)):
            return bytes(output)
        return output.encode("latin-1", errors="replace")

    def export(
        self,
        relatorio: Dict[str, Any],
        formato: str,
    ) -> Dict[str, Any]:
        normalized = (formato or "").strip().lower()
        if normalized == "pdf":
            payload = self.build_pdf_bytes(relatorio)
            extension = "pdf"
            media_type = "application/pdf"
        elif normalized == "csv":
            payload = self.build_csv_bytes(relatorio)
            extension = "csv"
            media_type = "text/csv; charset=utf-8"
        elif normalized == "excel":
            payload = self.build_excel_bytes(relatorio)
            extension = "xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise ValueError("Formato de exportação inválido. Use: pdf, csv ou excel.")

        filename = f"{self._build_filename_base(relatorio)}.{extension}"
        return {
            "payload": payload,
            "media_type": media_type,
            "filename": filename,
        }
